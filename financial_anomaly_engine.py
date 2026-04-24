"""
Financial Data Reporting & Anomaly Detection Engine
====================================================
Detects anomalies in regional financial data, scores severity,
and generates a professional Excel reporting dashboard.

Tools: Python, Pandas, OpenPyXL
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ── 1. LOAD ────────────────────────────────────────────────────
df = pd.read_csv("data/financial_raw_data.csv")
print(f"📂 Loaded {len(df)} records")

# ── 2. ANOMALY DETECTION LOGIC ─────────────────────────────────
flags = pd.DataFrame({"Record_ID": df["Record_ID"]})

# Check 1: Negative Revenue
flags["Negative_Revenue"] = df["Revenue"] < 0

# Check 2: Zero Revenue (but units sold > 0)
flags["Zero_Revenue"] = (df["Revenue"] == 0) & (df["Units_Sold"] > 0)

# Check 3: Cost exceeds Revenue
flags["Cost_Exceeds_Revenue"] = df["Cost"] > df["Revenue"]

# Check 4: Revenue spike — > mean + 3*std per product
rev_stats = df.groupby("Product")["Revenue"].agg(["mean","std"]).reset_index()
rev_stats.columns = ["Product","rev_mean","rev_std"]
df2 = df.merge(rev_stats, on="Product")
flags["Revenue_Spike"] = df2["Revenue"] > (df2["rev_mean"] + 3 * df2["rev_std"])

# Check 5: Units sold = 0 but revenue exists
flags["Units_Zero_Revenue_Exists"] = (df["Units_Sold"] == 0) & (df["Revenue"] > 0)

# Check 6: Missing Region
flags["Missing_Region"] = df["Region"].isna() | df["Region"].eq("")

# Check 7: Profit margin < -20% (severe loss)
margin = (df["Profit"] / df["Revenue"].replace(0, float("nan"))).fillna(0)
flags["Severe_Loss"] = margin < -0.20

# Severity scoring
weights = {
    "Negative_Revenue": 5,
    "Revenue_Spike": 4,
    "Cost_Exceeds_Revenue": 3,
    "Severe_Loss": 3,
    "Units_Zero_Revenue_Exists": 2,
    "Zero_Revenue": 2,
    "Missing_Region": 1,
}

bool_cols = list(weights.keys())
flags["Anomaly_Score"] = sum(flags[c].astype(int) * w for c, w in weights.items())
flags["Is_Anomaly"] = flags["Anomaly_Score"] > 0
flags["Severity"] = flags["Anomaly_Score"].apply(
    lambda s: "CRITICAL" if s >= 5 else ("HIGH" if s >= 3 else ("MEDIUM" if s > 0 else "CLEAN"))
)
flags["Flag_Count"] = flags[bool_cols].sum(axis=1)

# Merge
audit = df.merge(flags.drop(columns=["Record_ID"]), left_index=True, right_index=True)

# ── 3. SUMMARY STATS ──────────────────────────────────────────
total_revenue = df[(df["Revenue"] > 0) & (~flags["Revenue_Spike"])]["Revenue"].sum()
total_anomalies = flags["Is_Anomaly"].sum()

summary = {
    "Total Records": len(df),
    "Clean Records": (~flags["Is_Anomaly"]).sum(),
    "Anomalies Detected": int(total_anomalies),
    "Critical": (flags["Severity"] == "CRITICAL").sum(),
    "High": (flags["Severity"] == "HIGH").sum(),
    "Total Revenue (Clean)": round(total_revenue, 2),
}

region_summary = audit[audit["Region"].ne("") & audit["Region"].notna()].groupby("Region").agg(
    Total_Records=("Record_ID","count"),
    Total_Revenue=("Revenue", lambda x: x[x>0].sum()),
    Total_Profit=("Profit","sum"),
    Anomalies=("Is_Anomaly","sum"),
    Avg_Margin=("Profit", lambda x: (x.sum() / audit.loc[x.index,"Revenue"].replace(0,float("nan")).sum()) * 100)
).round(2).reset_index()

monthly = audit[audit["Revenue"] > 0].groupby("Month").agg(
    Revenue=("Revenue","sum"),
    Profit=("Profit","sum"),
    Anomalies=("Is_Anomaly","sum")
).round(2).reset_index()

product_summary = audit[audit["Revenue"] > 0].groupby("Product").agg(
    Revenue=("Revenue","sum"),
    Units=("Units_Sold","sum"),
    Anomalies=("Is_Anomaly","sum")
).round(2).reset_index()

issue_counts = flags[bool_cols].sum().reset_index()
issue_counts.columns = ["Issue_Type","Count"]
issue_counts = issue_counts[issue_counts["Count"] > 0].sort_values("Count", ascending=False)

print(f"\n📊 Summary:")
for k,v in summary.items(): print(f"   {k}: {v}")

# ── 4. EXCEL REPORT ───────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Colors
C_NAVY   = "1F3864"
C_BLUE   = "2E75B6"
C_RED    = "C00000"
C_ORANGE = "ED7D31"
C_GREEN  = "70AD47"
C_GOLD   = "C9A227"
C_WHITE  = "FFFFFF"
C_LIGHT  = "EEF4FF"
C_ALT    = "F5F9FF"

thin = Border(*[Side(style='thin', color="CCCCCC")] * 0,
              left=Side(style='thin', color="CCCCCC"),
              right=Side(style='thin', color="CCCCCC"),
              top=Side(style='thin', color="CCCCCC"),
              bottom=Side(style='thin', color="CCCCCC"))

def hdr(cell, bg=C_NAVY, fg=C_WHITE, sz=11, bold=True):
    cell.font = Font(bold=bold, color=fg, size=sz, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

def dat(cell, bg=C_WHITE, bold=False, color="222222", center=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = thin

# ════════════════════════
# SHEET 1: DASHBOARD
# ════════════════════════
ws1 = wb.create_sheet("📊 Dashboard")
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:K2")
ws1["A1"] = "📈  FINANCIAL DATA REPORTING & ANOMALY DETECTION DASHBOARD"
ws1["A1"].font = Font(bold=True, size=17, color=C_WHITE, name="Arial")
ws1["A1"].fill = PatternFill("solid", start_color=C_NAVY)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 8

ws1.merge_cells("A3:K3")
ws1["A3"] = f"Period: Jan 2023 – Dec 2023   |   Regions: North, South, East, West, Central   |   Products: A–E   |   Total Records: {len(df)}"
ws1["A3"].font = Font(italic=True, size=10, color="444444", name="Arial")
ws1["A3"].alignment = Alignment(horizontal="center")
ws1["A3"].fill = PatternFill("solid", start_color=C_LIGHT)

# KPI Cards
kpis = [
    ("Total Records", f"{summary['Total Records']:,}", C_BLUE),
    ("✅ Clean Records", f"{summary['Clean Records']:,}", C_GREEN),
    ("🚨 Anomalies", f"{summary['Anomalies Detected']:,}", C_RED),
    ("🔴 Critical", f"{summary['Critical']:,}", "8B0000"),
    ("💰 Clean Revenue", f"₹{summary['Total Revenue (Clean)']:,.0f}", "1D6A4A"),
]

ws1.row_dimensions[5].height = 18
ws1.row_dimensions[6].height = 46
ws1.row_dimensions[7].height = 8

kpi_start_cols = [1, 3, 5, 7, 9]
for idx, (label, value, color) in enumerate(kpis):
    c = kpi_start_cols[idx]
    cl = get_column_letter(c)
    cl2 = get_column_letter(c+1)
    ws1.merge_cells(f"{cl}5:{cl2}5")
    ws1.merge_cells(f"{cl}6:{cl2}6")
    ws1[f"{cl}5"] = label
    ws1[f"{cl}5"].font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
    ws1[f"{cl}5"].fill = PatternFill("solid", start_color=color)
    ws1[f"{cl}5"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"{cl}6"] = value
    ws1[f"{cl}6"].font = Font(bold=True, size=16, color=color, name="Arial")
    ws1[f"{cl}6"].fill = PatternFill("solid", start_color="F8F8F8")
    ws1[f"{cl}6"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"{cl}6"].border = thin

# Anomaly type table
row = 9
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"] = "Anomaly Type Breakdown"
hdr(ws1[f"A{row}"], bg=C_BLUE)
row += 1
for col, h in zip("ABCD", ["#", "Anomaly Type", "Count", "% of Total"]):
    ws1[f"{col}{row}"] = h
    hdr(ws1[f"{col}{row}"], bg=C_BLUE)
for i, (_, r) in enumerate(issue_counts.iterrows()):
    row += 1
    bg = C_ALT if i % 2 == 0 else C_WHITE
    pct = f"{r['Count']/len(df)*100:.1f}%"
    for col, val in zip("ABCD", [i+1, r["Issue_Type"].replace("_"," "), int(r["Count"]), pct]):
        ws1[f"{col}{row}"] = val
        dat(ws1[f"{col}{row}"], bg=bg, center=(col in "ACD"))

# Region table
row2 = 9
ws1.merge_cells(f"F{row2}:K{row2}")
ws1[f"F{row2}"] = "Region Performance Summary"
hdr(ws1[f"F{row2}"], bg=C_BLUE)
row2 += 1
for col, h in zip("FGHIJK", ["Region","Records","Revenue (₹)","Profit (₹)","Anomalies","Avg Margin %"]):
    ws1[f"{col}{row2}"] = h
    hdr(ws1[f"{col}{row2}"], bg=C_BLUE)
for i, (_, r) in enumerate(region_summary.iterrows()):
    row2 += 1
    bg = C_ALT if i % 2 == 0 else C_WHITE
    vals = [r["Region"], int(r["Total_Records"]), f"₹{r['Total_Revenue']:,.0f}",
            f"₹{r['Total_Profit']:,.0f}", int(r["Anomalies"]), f"{r['Avg_Margin']:.1f}%"]
    for col, val in zip("FGHIJK", vals):
        ws1[f"{col}{row2}"] = val
        dat(ws1[f"{col}{row2}"], bg=bg, center=(col != "F"))

ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 2
for col in "FGHIJK":
    ws1.column_dimensions[col].width = 15

# ════════════════════════
# SHEET 2: ANOMALY LOG
# ════════════════════════
ws2 = wb.create_sheet("🚨 Anomaly Log")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"

anomalies_df = audit[audit["Is_Anomaly"]].sort_values("Anomaly_Score", ascending=False)
show_cols = ["Record_ID","Month","Region","Product","Channel",
             "Units_Sold","Revenue","Cost","Profit","Anomaly_Score","Severity"] + bool_cols

ws2.merge_cells(f"A1:{get_column_letter(len(show_cols))}1")
ws2["A1"] = f"🚨 ANOMALY LOG — {len(anomalies_df)} Records Flagged for Investigation"
ws2["A1"].font = Font(bold=True, size=12, color=C_WHITE, name="Arial")
ws2["A1"].fill = PatternFill("solid", start_color=C_RED)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 25

for ci, h in enumerate(show_cols, 1):
    hdr(ws2.cell(row=2, column=ci, value=h.replace("_"," ")), bg=C_RED)

sev_colors = {"CRITICAL": "8B0000", "HIGH": C_RED, "MEDIUM": C_ORANGE}

for ri, (_, row_data) in enumerate(anomalies_df[show_cols].iterrows(), 3):
    sev = row_data["Severity"]
    bg = {"CRITICAL": "FFE8E8", "HIGH": "FFF0F0", "MEDIUM": "FFF8F0"}.get(sev, C_WHITE)
    for ci, (col, val) in enumerate(row_data.items(), 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        if col == "Severity":
            dat(cell, bg=sev_colors.get(sev, C_WHITE), bold=True, center=True)
            cell.font = Font(bold=True, color=C_WHITE, name="Arial", size=10)
        elif col in bool_cols and val:
            dat(cell, bg="FFD6D6", center=True)
            cell.value = "⚠ FLAG"
            cell.font = Font(bold=True, color=C_RED, name="Arial", size=9)
        elif col in bool_cols:
            dat(cell, bg=bg, center=True)
            cell.value = "✓"
            cell.font = Font(color=C_GREEN, name="Arial", size=10)
        else:
            dat(cell, bg=bg, center=(col in ["Anomaly_Score","Units_Sold","Revenue","Cost","Profit"]))

for ci in range(1, len(show_cols)+1):
    ws2.column_dimensions[get_column_letter(ci)].width = 15

# ════════════════════════
# SHEET 3: MONTHLY TREND
# ════════════════════════
ws3 = wb.create_sheet("📈 Monthly Trend")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:E1")
ws3["A1"] = "Monthly Revenue, Profit & Anomaly Trend — Jan to Dec 2023"
hdr(ws3["A1"], bg=C_NAVY, sz=12)
ws3.row_dimensions[1].height = 25

for ci, h in enumerate(["Month","Revenue (₹)","Profit (₹)","Anomalies","Profit Margin %"], 1):
    hdr(ws3.cell(row=2, column=ci), bg=C_BLUE)

for ri, (_, r) in enumerate(monthly.iterrows(), 3):
    bg = C_ALT if ri % 2 == 0 else C_WHITE
    margin = f"{r['Profit']/r['Revenue']*100:.1f}%" if r["Revenue"] > 0 else "N/A"
    for ci, val in enumerate([r["Month"], f"₹{r['Revenue']:,.0f}",
                               f"₹{r['Profit']:,.0f}", int(r["Anomalies"]), margin], 1):
        dat(ws3.cell(row=ri, column=ci, value=val), bg=bg, center=(ci > 1))

# Bar chart — Revenue by month
chart_data_start = 3
chart_data_end = 3 + len(monthly) - 1

bar = BarChart()
bar.type = "col"
bar.title = "Monthly Revenue"
bar.y_axis.title = "Revenue (₹)"
bar.x_axis.title = "Month"
bar.width = 20
bar.height = 12
data_ref = Reference(ws3, min_col=2, min_row=2, max_row=chart_data_end)
cats = Reference(ws3, min_col=1, min_row=3, max_row=chart_data_end)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = "2E75B6"
ws3.add_chart(bar, "G2")

for ci, w in enumerate([14,16,16,12,14], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ════════════════════════
# SHEET 4: CLEAN DATA
# ════════════════════════
ws4 = wb.create_sheet("✅ Clean Data")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A2"

clean_df = audit[~audit["Is_Anomaly"]].copy()
clean_cols = ["Record_ID","Month","Region","Product","Channel","Units_Sold","Unit_Price","Revenue","Cost","Profit"]

ws4.merge_cells(f"A1:{get_column_letter(len(clean_cols))}1")
ws4["A1"] = f"✅ CLEAN RECORDS — Passed All Checks  ({len(clean_df)} records)"
ws4["A1"].font = Font(bold=True, size=12, color=C_WHITE, name="Arial")
ws4["A1"].fill = PatternFill("solid", start_color="375623")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")

for ci, h in enumerate(clean_cols, 1):
    hdr(ws4.cell(row=2, column=ci), bg="375623")

for ri, (_, r) in enumerate(clean_df[clean_cols].iterrows(), 3):
    bg = C_ALT if ri % 2 == 0 else C_WHITE
    for ci, val in enumerate(r.values, 1):
        dat(ws4.cell(row=ri, column=ci, value=val), bg=bg)

for ci in range(1, len(clean_cols)+1):
    ws4.column_dimensions[get_column_letter(ci)].width = 15

# ════════════════════════
# SHEET 5: SQL REFERENCE
# ════════════════════════
ws5 = wb.create_sheet("🗄 SQL Queries")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:B1")
ws5["A1"] = "SQL Reconciliation & Anomaly Detection Queries (MySQL)"
hdr(ws5["A1"], bg=C_NAVY, sz=13)
ws5.row_dimensions[1].height = 28

queries = [
    ("1. Revenue by Region — Reconciliation",
     "SELECT region,\n  COUNT(*) AS records,\n  SUM(revenue) AS total_revenue,\n  SUM(profit) AS total_profit,\n  ROUND(SUM(profit)/NULLIF(SUM(revenue),0)*100, 2) AS margin_pct\nFROM financial_data\nWHERE region IS NOT NULL AND region != ''\nGROUP BY region\nORDER BY total_revenue DESC;"),
    ("2. Detect Negative Revenue",
     "SELECT record_id, month, region, product, revenue\nFROM financial_data\nWHERE revenue < 0\nORDER BY revenue ASC;"),
    ("3. Cost Exceeds Revenue (Loss Anomaly)",
     "SELECT record_id, month, region, product,\n  revenue, cost,\n  ROUND(cost - revenue, 2) AS excess_cost\nFROM financial_data\nWHERE cost > revenue\nORDER BY excess_cost DESC;"),
    ("4. Revenue Spike Detection (> 3 Std Dev)",
     "SELECT f.record_id, f.region, f.product, f.revenue,\n  s.avg_rev, s.std_rev,\n  ROUND((f.revenue - s.avg_rev) / NULLIF(s.std_rev, 0), 2) AS z_score\nFROM financial_data f\nJOIN (\n  SELECT product,\n    AVG(revenue) AS avg_rev,\n    STDDEV(revenue) AS std_rev\n  FROM financial_data\n  GROUP BY product\n) s ON f.product = s.product\nWHERE f.revenue > s.avg_rev + 3 * s.std_rev\nORDER BY z_score DESC;"),
    ("5. Units = 0 but Revenue > 0",
     "SELECT record_id, month, region, product, units_sold, revenue\nFROM financial_data\nWHERE units_sold = 0 AND revenue > 0;"),
    ("6. Monthly Revenue Trend",
     "SELECT month,\n  SUM(revenue) AS total_revenue,\n  SUM(profit) AS total_profit,\n  COUNT(*) AS transactions,\n  ROUND(SUM(profit)/NULLIF(SUM(revenue),0)*100, 2) AS margin_pct\nFROM financial_data\nGROUP BY month\nORDER BY month;"),
    ("7. Product-wise Anomaly Count",
     "SELECT product,\n  COUNT(*) AS total_records,\n  SUM(CASE WHEN revenue < 0 THEN 1 ELSE 0 END) AS neg_revenue,\n  SUM(CASE WHEN cost > revenue THEN 1 ELSE 0 END) AS cost_exceeds,\n  SUM(CASE WHEN units_sold = 0 AND revenue > 0 THEN 1 ELSE 0 END) AS units_zero\nFROM financial_data\nGROUP BY product\nORDER BY (neg_revenue + cost_exceeds + units_zero) DESC;"),
]

row = 3
for title, query in queries:
    ws5.merge_cells(f"A{row}:B{row}")
    ws5[f"A{row}"] = title
    ws5[f"A{row}"].font = Font(bold=True, size=11, color=C_WHITE, name="Arial")
    ws5[f"A{row}"].fill = PatternFill("solid", start_color=C_BLUE)
    ws5[f"A{row}"].alignment = Alignment(vertical="center")
    ws5.row_dimensions[row].height = 20
    row += 1
    lines = query.count("\n") + 1
    ws5.merge_cells(f"A{row}:B{row+lines}")
    ws5[f"A{row}"] = query
    ws5[f"A{row}"].font = Font(name="Courier New", size=10, color="1F3864")
    ws5[f"A{row}"].fill = PatternFill("solid", start_color="F0F4FF")
    ws5[f"A{row}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws5.row_dimensions[row].height = 14 * lines
    row += lines + 2

ws5.column_dimensions["A"].width = 85
ws5.column_dimensions["B"].width = 1

# ── SAVE ──────────────────────────────────────────────────────
os.makedirs("excel", exist_ok=True)
path = "excel/Financial_Anomaly_Report.xlsx"
wb.save(path)
print(f"\n✅ Excel saved → {path}")
print(f"   Sheets: Dashboard | Anomaly Log | Monthly Trend | Clean Data | SQL Queries")
